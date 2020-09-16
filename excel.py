import openpyxl
import platform
import random
import string

from openpyxl import load_workbook
import extra_functions

from openpyxl.utils.exceptions import IllegalCharacterError

excel_product_image_url = []
excel_product_names = []
work_sheet_index = 2
table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
cell_index = 0

d = {}


def write_product_code_to_excel(image_file_names, image_names_cell_letter):
    global work_sheet_index

    work_sheet_index = 2

    wb = openpyxl.load_workbook(table_location)
    ws = wb.active

    for i in range(len(image_file_names)):
        print('wbind: ', work_sheet_index)
        product_brand_key = extra_functions.value_key(image_names_cell_letter, work_sheet_index)
        ws[product_brand_key] = image_file_names[i]

        work_sheet_index += 1

    wb.save(table_location)


def match_key_value(key_cell, value_cell):
    global d
    global work_sheet_index
    global cell_index
    cells = ['CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI',
             'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ',
             'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ',
             'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DO', 'DI',
             'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR',
             'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ',
             'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EO', 'EI',
             'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER',
             'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ',
             'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FO', 'FI',
             'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR',
             'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ',
             'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GO', 'GI',
             'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GR',
             'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ',
             'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG', 'HO', 'HI',
             'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ', 'HR',
             'HS', 'HT', 'HU', 'HV', 'HW', 'HX', 'HY', 'HZ',
             'IA', 'IB', 'IC', 'ID', 'IE', 'IF', 'IG', 'IO', 'II',
             'IJ', 'IK', 'IL', 'IM', 'IN', 'IO', 'IP', 'IQ', 'IR',
             'IS', 'IT', 'IU', 'IV', 'IW', 'IX', 'IY', 'IZ',
             'JA', 'JB', 'JC', 'JD', 'JE', 'JF', 'JG', 'JO', 'JI',
             'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP', 'JQ', 'JR',
             'JS', 'JT', 'JU', 'JV', 'JW', 'JX', 'JY', 'JZ',
             'KA', 'KB', 'KC', 'KD', 'KE', 'KF', 'KG', 'KO', 'KI',
             'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', 'KR',
             'KS', 'KT', 'KU', 'KV', 'KW', 'KX', 'KY', 'KZ',
             'LA', 'LB', 'LC', 'LD', 'LE', 'LF', 'LG', 'LO', 'LI',
             'LJ', 'LK', 'LL', 'LM', 'LN', 'LO', 'LP', 'LQ', 'LR',
             'LS', 'LT', 'LU', 'LV', 'LW', 'LX', 'LY', 'LZ',
             'MA', 'MB', 'MC', 'MD', 'ME', 'MF', 'MG', 'MO', 'MI',
             'MJ', 'MK', 'ML', 'MM', 'MN', 'MO', 'MP', 'MQ', 'MR',
             'MS', 'MT', 'MU', 'MV', 'MW', 'MX', 'MY', 'MZ',
             'NA', 'NB', 'NC', 'ND', 'NE', 'NF', 'NG', 'NO', 'NI',
             'NJ', 'NK', 'NL', 'NM', 'NN', 'NO', 'NP', 'NQ', 'NR',
             'NS', 'NT', 'NU', 'NV', 'NW', 'NX', 'NY', 'NZ',
             'OA', 'OB', 'OC', 'OD', 'OE', 'OF', 'OG', 'OO', 'OI',
             'OJ', 'OK', 'OL', 'OM', 'ON', 'OO', 'OP', 'OQ', 'OR',
             'OS', 'OT', 'OU', 'OV', 'OW', 'OX', 'OY', 'OZ',
             'PA', 'PB', 'PC', 'PD', 'PE', 'PF', 'PG', 'PO', 'PI',
             'PJ', 'PK', 'PL', 'PM', 'PN', 'PO', 'PP', 'PQ', 'PR',
             'PS', 'PT', 'PU', 'PV', 'PW', 'PX', 'PY', 'PZ',
             'QA', 'QB', 'QC', 'QD', 'QE', 'QF', 'QG', 'QO', 'QI',
             'QJ', 'QK', 'QL', 'QM', 'QN', 'QO', 'QP', 'QQ', 'QR',
             'QS', 'QT', 'QU', 'QV', 'QW', 'QX', 'QY', 'QZ',
             'RA', 'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RO', 'RI',
             'RJ', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RQ', 'RR',
             'RS', 'RT', 'RU', 'RV', 'RW', 'RX', 'RY', 'RZ',

             ]
    wb = load_workbook(table_location)  # Work Book
    ws = wb.active  # Work Sheet
    column = ws['A']

    work_sheet_index = 2

    for i in range(len(column)):
        key = ws[key_cell + str(work_sheet_index)].value
        cell = cells[cell_index]
        value = ws[value_cell + str(work_sheet_index)].value

        if key not in d:
            print('cell index: NEW', cell_index)
            print('work sheet index: ', work_sheet_index)
            d.update({key: cell})
            ws[cell + '1'] = key
            key_to_look_in_dict = ws[key_cell + str(work_sheet_index)].value
            # print(f'Key cell is: {key_cell}{str(work_sheet_index)} we to look in the dict: ', key_to_look_in_dict)
            print(f'in cell {d[key_to_look_in_dict]}{str(work_sheet_index)} we put the value: '
                  f'{ws[value_cell + str(work_sheet_index)].value}  ')
            value = ws[value_cell + str(work_sheet_index)].value
            ws[cell + str(work_sheet_index)] = value
            cell_index += 1
            work_sheet_index += 1
        else:
            print('cell index: EXISTING', cell_index)
            print('work sheet index: ', work_sheet_index)
            key_to_look_in_dict = ws[key_cell + str(work_sheet_index)].value
            # print(f'Key cell is: {key_cell}{str(work_sheet_index)} we to look in the dict: ', key_to_look_in_dict)
            print(f'in cell {d[key_to_look_in_dict]}{str(work_sheet_index)} we put the value: '
                  f'{ws[value_cell + str(work_sheet_index)].value}  ')
            if ws[d[key_to_look_in_dict] + str(work_sheet_index)].value is None:
                print('work sheet index: ', work_sheet_index)
                ws[d[key_to_look_in_dict] + str(work_sheet_index)] = ws[
                    value_cell + str(work_sheet_index)].value
            else:
                print('error')
            work_sheet_index += 1
    wb.save(table_location)
    return cell_index


def get_all_the_rows_from_column(cell_letter):
    wb = load_workbook(table_location)  # Work Book
    ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
    column = ws[cell_letter]  # Column
    column_list = [column[x].value for x in range(1, len(column))]
    return column_list


def get_work_sheet_index():
    global work_sheet_index
    wb = load_workbook(table_location)  # Work Book
    ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
    column = ws['A']  # Column
    work_sheet_index = len(column) + 1
    print('Worksheet index: ', work_sheet_index)
